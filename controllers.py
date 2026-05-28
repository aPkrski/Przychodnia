from datetime import date, datetime

import pandas as pd
from sqlalchemy import cast, Date, func, or_, select

from database import SessionLocal, init_db
from models import Clinic, Invoice, Location, Payroll, Revenue


class AppController:
    def __init__(self):
        init_db()

    def _session(self):
        return SessionLocal()

    def get_locations(self):
        with self._session() as session:
            return session.scalars(select(Location).order_by(Location.name)).all()

    def get_clinics(self, location_id=None):
        with self._session() as session:
            query = select(Clinic)
            if location_id is not None:
                query = query.where(Clinic.location_id == location_id)
            return session.scalars(query.order_by(Clinic.number, Clinic.name)).all()

    def get_invoices(self, clinic_id, start_date=None, end_date=None, search_text=None):
        return self._get_records(Invoice, clinic_id, start_date, end_date, search_text)

    def get_payrolls(self, clinic_id, start_date=None, end_date=None, search_text=None):
        return self._get_records(Payroll, clinic_id, start_date, end_date, search_text)

    def get_revenues(self, clinic_id, start_date=None, end_date=None, search_text=None):
        return self._get_records(Revenue, clinic_id, start_date, end_date, search_text)

    def _get_records(self, model, clinic_id, start_date, end_date, search_text):
        with self._session() as session:
            query = select(model).where(model.clinic_id == clinic_id)
            if start_date:
                query = query.where(model.date >= start_date)
            if end_date:
                query = query.where(model.date <= end_date)
            if search_text:
                search = f"%{search_text}%"
                if model is Invoice:
                    query = query.where(or_(model.number.ilike(search), model.item.ilike(search)))
                elif model is Payroll:
                    query = query.where(model.employee.ilike(search))
                elif model is Revenue:
                    query = query.where(model.company.ilike(search))
            return session.scalars(query.order_by(model.date.desc())).all()

    def _add_record(self, model, clinic_id, values):
        with self._session() as session:
            record = model(clinic_id=clinic_id, **values)
            session.add(record)
            session.commit()
            session.refresh(record)
            return record

    def add_invoice(self, clinic_id, number, item, net_amount, gross_amount, date, category="", company_name=""):
        return self._add_record(Invoice, clinic_id, {
            "number": number,
            "item": item,
            "net_amount": net_amount,
            "gross_amount": gross_amount,
            "date": date,
            "category": category or "",
            "company_name": company_name or "",
        })

    def add_payroll(self, clinic_id, employee, period, amount, date):
        return self._add_record(Payroll, clinic_id, {
            "employee": employee,
            "period": period,
            "amount": amount,
            "date": date,
        })

    def add_revenue(self, clinic_id, company, period, amount, date):
        return self._add_record(Revenue, clinic_id, {
            "company": company,
            "period": period,
            "amount": amount,
            "date": date,
        })

    def update_record(self, model, record_id, data):
        with self._session() as session:
            record = session.get(model, record_id)
            if record is None:
                return None
            for key, value in data.items():
                setattr(record, key, value)
            session.commit()
            session.refresh(record)
            return record

    def delete_record(self, model, record_id):
        with self._session() as session:
            record = session.get(model, record_id)
            if record:
                session.delete(record)
                session.commit()
                return True
            return False

    def get_financial_summary(self, location_id=None, clinic_id=None, start_date=None, end_date=None):
        with self._session() as session:
            invoice_stmt = select(func.coalesce(func.sum(Invoice.net_amount), 0), func.coalesce(func.sum(Invoice.gross_amount), 0))
            payroll_stmt = select(func.coalesce(func.sum(Payroll.amount), 0))
            revenue_stmt = select(func.coalesce(func.sum(Revenue.amount), 0))
            if clinic_id:
                invoice_stmt = invoice_stmt.where(Invoice.clinic_id == clinic_id)
                payroll_stmt = payroll_stmt.where(Payroll.clinic_id == clinic_id)
                revenue_stmt = revenue_stmt.where(Revenue.clinic_id == clinic_id)
            elif location_id:
                clinic_ids = session.scalars(select(Clinic.id).where(Clinic.location_id == location_id)).all()
                invoice_stmt = invoice_stmt.where(Invoice.clinic_id.in_(clinic_ids))
                payroll_stmt = payroll_stmt.where(Payroll.clinic_id.in_(clinic_ids))
                revenue_stmt = revenue_stmt.where(Revenue.clinic_id.in_(clinic_ids))
            if start_date:
                invoice_stmt = invoice_stmt.where(Invoice.date >= start_date)
                payroll_stmt = payroll_stmt.where(Payroll.date >= start_date)
                revenue_stmt = revenue_stmt.where(Revenue.date >= start_date)
            if end_date:
                invoice_stmt = invoice_stmt.where(Invoice.date <= end_date)
                payroll_stmt = payroll_stmt.where(Payroll.date <= end_date)
                revenue_stmt = revenue_stmt.where(Revenue.date <= end_date)
            net_sum, gross_sum = session.execute(invoice_stmt).one()
            payroll_sum = session.execute(payroll_stmt).scalar_one()
            revenue_sum = session.execute(revenue_stmt).scalar_one()
            profit = revenue_sum - payroll_sum - net_sum
            return {
                "invoice_net": float(net_sum),
                "invoice_gross": float(gross_sum),
                "payroll": float(payroll_sum),
                "revenue": float(revenue_sum),
                "profit": float(profit),
            }

    def export_to_excel(self, items, filename):
        df = pd.DataFrame(items)
        if df.empty:
            df = pd.DataFrame(columns=["ID", "Clinic", "Type", "Value"])
        df.to_excel(filename, index=False, engine="openpyxl")

    def export_clinic_summary(self, clinic_id, filename, start_date=None, end_date=None):
        invoices = self.get_invoices(clinic_id, start_date=start_date, end_date=end_date)
        payrolls = self.get_payrolls(clinic_id, start_date=start_date, end_date=end_date)
        revenues = self.get_revenues(clinic_id, start_date=start_date, end_date=end_date)

        inv_rows = []
        for r in invoices:
            inv_rows.append({
                "ID": r.id,
                "Numer faktury": r.number,
                "Pozycja": r.item,
                "Cena netto": float(r.net_amount),
                "Cena brutto": float(r.gross_amount),
                "Data": r.date,
            })

        pay_rows = []
        for p in payrolls:
            pay_rows.append({
                "ID": p.id,
                "Pracownik": p.employee,
                "Miesiąc": p.period,
                "Kwota": float(p.amount),
                "Data": p.date,
            })

        rev_rows = []
        for r in revenues:
            rev_rows.append({
                "ID": r.id,
                "Firma": r.company,
                "Miesiąc": r.period,
                "Kwota": float(r.amount),
                "Data": r.date,
            })

        summary = self.get_financial_summary(clinic_id=clinic_id)
        summary_rows = [
            {"Kategoria": "Suma faktur netto", "Kwota": summary["invoice_net"]},
            {"Kategoria": "Suma faktur brutto", "Kwota": summary["invoice_gross"]},
            {"Kategoria": "Suma wynagrodzeń", "Kwota": summary["payroll"]},
            {"Kategoria": "Suma przychodów", "Kwota": summary["revenue"]},
            {"Kategoria": "Wynik końcowy", "Kwota": summary["profit"]},
        ]

        df_inv = pd.DataFrame(inv_rows)
        df_pay = pd.DataFrame(pay_rows)
        df_rev = pd.DataFrame(rev_rows)
        df_summary = pd.DataFrame(summary_rows)

        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df_summary.to_excel(writer, sheet_name="Podsumowanie", index=False)
            df_inv.to_excel(writer, sheet_name="Faktury", index=False)
            df_pay.to_excel(writer, sheet_name="Wynagrodzenia", index=False)
            df_rev.to_excel(writer, sheet_name="Przychody", index=False)
            meta = pd.DataFrame([{
                "Wygenerowano": datetime.now().isoformat(),
                "ZakresDat": f"{start_date} do {end_date}" if start_date or end_date else "wszystkie"
            }])
            meta.to_excel(writer, sheet_name="Metadane", index=False)

    def export_report(self, filename, sheets: dict, image_path: str = None, metadata: dict = None):
        """
        sheets: dict of sheet_name -> pandas.DataFrame
        image_path: optional path to an image to insert into the first sheet
        metadata: dict of metadata values to add to Metadata sheet
        """
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            for name, df in sheets.items():
                # ensure dataframe
                if df is None or (hasattr(df, 'empty') and df.empty):
                    pd.DataFrame().to_excel(writer, sheet_name=name, index=False)
                else:
                    df.to_excel(writer, sheet_name=name, index=False)
            meta_rows = [metadata] if metadata else []
            meta_rows.append({"Wygenerowano": datetime.now().isoformat()})
            pd.DataFrame(meta_rows).to_excel(writer, sheet_name="Metadane", index=False)
            writer.save()
        # insert image using openpyxl if provided
        try:
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image as OpenpyxlImage
            wb = load_workbook(filename)
            # set number formats for numeric currency columns
            for sheet_name in sheets.keys():
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    # read headers
                    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                    currency_cols = [i for i, h in enumerate(headers) if isinstance(h, str) and ("Cena" in h or "Kwota" in h or "Brutto" in h or "netto" in h or "brutto" in h.lower())]
                    if currency_cols:
                        for row in ws.iter_rows(min_row=2):
                            for col_idx in currency_cols:
                                try:
                                    cell = row[col_idx]
                                    # if cell contains numeric value, set number format
                                    if cell.value is not None:
                                        try:
                                            cell.value = float(cell.value)
                                            cell.number_format = '#,##0.00 "zł"'
                                        except Exception:
                                            pass
                                except Exception:
                                    pass
            # insert image into first sheet if provided
            if image_path:
                try:
                    first_sheet = list(sheets.keys())[0]
                    ws = wb[first_sheet]
                    img = OpenpyxlImage(image_path)
                    img.anchor = 'A1'
                    ws.add_image(img)
                except Exception:
                    pass
            wb.save(filename)
        except Exception:
            pass

    def get_location_by_name(self, location_name):
        with self._session() as session:
            return session.scalar(select(Location).where(Location.name == location_name))

    def get_clinic_by_name(self, location_id, clinic_name):
        with self._session() as session:
            return session.scalar(select(Clinic).where(Clinic.location_id == location_id, Clinic.name == clinic_name))
